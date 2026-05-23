"""Regenerate synthetic_workbook.xlsx — a sanitized DKU master-workbook
fixture for parser tests (in this repo + lifted into vfc per prism plan
0039).

The shape mirrors the real `留本基金动态资产配置情况.xlsx` precisely
enough to exercise the parser's main paths:

    - title regex (`YYYY年MM月DD日基金会留本基金...`)
    - header row with `持有标的` substring matcher
    - the 4-bucket asset-class column layout (Growth / FI / Inflation /
      Diversifiers — column-position-encoded)
    - multi-lot funds (same fund, multiple subscription dates)
    - SUBTOTAL_MARKERS (`主动配置小计`, `留本基金总额`) ending the lot block
    - bottom summary block headed by `类别`
    - 备注 free-text on a redemption row (notes_parser pattern)
    - 8 funds spread across the 3 identifier tiers (Tier 1/2/3 per
      plan 0039 lazy-mint)
    - 2 snapshot sheets so cross-snapshot tests work

Sanitization vs the real workbook:

    - Costs are round (100k / 500k / 1M)
    - NAVs are clean 4-decimal numbers (1.0500 / 1.0823 etc.)
    - No external-workbook references; no structured-table formulas —
      every cell is a literal value, openpyxl can read it without
      `data_only=True`
    - The deferred sheets (`市值动态`, `收益回撤`, `债权比较`) are
      omitted; if the parser ever needs them, scaffold separately
    - 备注 text is generic ("于YYYY年M月D日全部赎回, 赎回净值 X.XXXX") —
      no DKU-internal operator names or counterparty details

Run:  uv run python tests/fixtures/master_record/_generate_synthetic.py
Output: tests/fixtures/master_record/synthetic_workbook.xlsx
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

OUT = Path(__file__).parent / "synthetic_workbook.xlsx"


# Tier 1 — explicit `(NNNNNN)` in the workbook (公募 funds, AMAC-public)
# Tier 2 — code-in-parens via dkup `fund_aliases.parquet` curation
# Tier 3 — name-only 私募 funds outside AMAC public coverage
FUNDS = [
    # (display_string, asset_class_column, tier)
    ("华夏纯债债券A\n(000015)",       "bucket_fi",          1),
    ("易方达裕祥回报债券\n(002351)",   "bucket_diversified", 1),
    ("华安黄金易(ETF联接)A\n(000216)", "bucket_inflation",   1),
    # Tier 2 — fund_aliases.parquet bridges name → code (no parens shown
    # in this snapshot, but the alias table maps it; for the synthetic
    # we treat it the same as Tier 3 from the parser's POV since the
    # parser sees no parens code on this row).
    ("海富通安颐 收益混合A",            "bucket_growth",      2),
    # Tier 3 — private, AMAC-invisible
    ("睿远基金睿见1号",                 "bucket_growth",      3),
    ("禾瑞十号",                       "bucket_fi",          3),
    ("禾禧五号",                       "bucket_fi",          3),
    ("宁泉致远18号",                   "bucket_diversified", 3),
]

BUCKET_COL = {
    "bucket_growth":      "E",  # col 5  投资增长类(30%)
    "bucket_fi":          "F",  # col 6  投资固定收益类(20%)
    "bucket_inflation":   "G",  # col 7  投资通胀敏感类(10-30%)
    "bucket_diversified": "H",  # col 8  投资分散投资类(0-30%)
}

# Multi-lot setup — fund_name → [(subscription_date, cost, nav, units)]
# These per-snapshot maps are used to build both 20260331 and 20260430
# sheets. NAV bumps slightly between snapshots so reconciliation tests
# can verify growth math without contrived numbers.
LOTS_20260331 = {
    "华夏纯债债券A\n(000015)":       [(date(2023, 1, 20), 500_000, 1.0823, None)],
    "易方达裕祥回报债券\n(002351)":   [(date(2021, 2,  1), 500_000, 1.1542, None)],
    "华安黄金易(ETF联接)A\n(000216)": [(date(2023, 3, 30), 400_000, 1.6700, None)],
    "海富通安颐 收益混合A":            [(date(2024, 6, 15), 1_000_000, 1.1234, None)],
    "睿远基金睿见1号":                 [(date(2025, 2, 6), 3_000_000, 1.0500, 3_000_000.00),
                                       (date(2025, 6, 5), 1_000_000, 1.0500, 1_000_000.00)],
    "禾瑞十号":                       [(date(2025, 2, 11), 1_000_000, 1.0200, 1_000_000.00),
                                       (date(2025, 6, 11), 2_000_000, 1.0200, 2_000_000.00)],
    "禾禧五号":                       [(date(2025, 2, 11), 1_000_000, 1.0300, 1_000_000.00)],
    "宁泉致远18号":                   [(date(2025, 2, 21), 1_000_000, 1.0150, 1_000_000.00)],
}

LOTS_20260430 = {
    "华夏纯债债券A\n(000015)":       [(date(2023, 1, 20), 500_000, 1.0876, None)],
    "易方达裕祥回报债券\n(002351)":   [(date(2021, 2,  1), 500_000, 1.1601, None)],
    "华安黄金易(ETF联接)A\n(000216)": [(date(2023, 3, 30), 400_000, 1.6850, None)],
    "海富通安颐 收益混合A":            [(date(2024, 6, 15), 1_000_000, 1.1289, None)],
    "睿远基金睿见1号":                 [(date(2025, 2, 6), 3_000_000, 1.0612, 3_000_000.00),
                                       (date(2025, 6, 5), 1_000_000, 1.0612, 1_000_000.00),
                                       (date(2025, 8, 4), 1_500_000, 1.0612, 1_500_000.00)],
    "禾瑞十号":                       [(date(2025, 2, 11), 1_000_000, 1.0287, 1_000_000.00),
                                       (date(2025, 6, 11), 2_000_000, 1.0287, 2_000_000.00),
                                       (date(2025, 8, 12), 2_000_000, 1.0287, 2_000_000.00)],
    # 禾禧五号 redeemed 2026-04-18 — the lot row stays (with original
    # subscription date + cost so the parser can resolve lot identity)
    # but units / mv collapse to 0 and the 备注 column carries the
    # redemption text for notes_parser to extract a redemption event.
    "禾禧五号":                       [(date(2025, 2, 11), 1_000_000, 1.0345, 0.0)],
    "宁泉致远18号":                   [(date(2025, 2, 21), 1_000_000, 1.0218, 1_000_000.00)],
}

REDEMPTION_NOTE = "于2026年4月18日全部赎回, 赎回净值 1.0345"


def _bucket_col_for(fund_display: str, snapshot_lots: dict) -> str:
    for f, b, _ in FUNDS:
        if f == fund_display:
            return BUCKET_COL[b]
    raise KeyError(fund_display)


def _write_snapshot(ws: Worksheet, snap_date: date, lots_by_fund: dict) -> None:
    """Write one snapshot sheet matching the real workbook's layout."""

    # Row 1 — title (parser regex needs `YYYY年M月D日...基金会留本基金`)
    ws["A1"] = f"{snap_date.year}年{snap_date.month}月{snap_date.day}日基金会留本基金资产配置情况"

    # Row 2 — header
    ws["A2"] = "持有标的"
    ws["B2"] = "投资时间"
    ws["C2"] = "投资成本"
    ws["D2"] = "占比"
    ws["E2"] = "投资增长类\n(30%)"
    ws["F2"] = "投资固定收益类\n(20%)"
    ws["G2"] = "投资通胀敏感类\n(10-30%)"
    ws["H2"] = "投资分散投资类\n(0-30%)"
    ws["I2"] = "标的价值"
    ws["J2"] = "单位净值"
    ws["K2"] = "份额"
    ws["L2"] = "备注"

    # Compute total cost for the % column
    total_cost = sum(cost for lots in lots_by_fund.values() for (_, cost, _, _) in lots)

    # Rows 3+ — lots
    r = 3
    for fund_display, _, _ in FUNDS:
        if fund_display not in lots_by_fund:
            continue
        bucket_col = _bucket_col_for(fund_display, lots_by_fund)
        for sub_date, cost, nav, units in lots_by_fund[fund_display]:
            ws[f"A{r}"] = fund_display
            ws[f"B{r}"] = sub_date
            ws[f"C{r}"] = cost
            ws[f"D{r}"] = cost / total_cost
            # units==0 ⇒ redeemed lot in this snapshot. Keep the lot
            # row visible (so the parser resolves lot identity) but
            # zero out the mv columns and stash the redemption text in
            # 备注 — that's the notes_parser entry point.
            is_redeemed = units == 0
            mv = 0.0 if is_redeemed else cost * nav
            ws[f"{bucket_col}{r}"] = mv
            ws[f"I{r}"] = mv
            ws[f"J{r}"] = nav
            if units is not None:
                ws[f"K{r}"] = units
            if is_redeemed:
                ws[f"L{r}"] = REDEMPTION_NOTE
            r += 1

    # 主动配置小计 — terminates the lot block
    subtotal_row = r
    ws[f"A{subtotal_row}"] = "主动配置小计"
    ws[f"C{subtotal_row}"] = total_cost
    r += 1

    # 现金产品小计 — small cash sleeve (sanitized)
    ws[f"A{r}"] = "现金产品小计"
    ws[f"C{r}"] = 200_000
    r += 1

    # 留本基金总额
    ws[f"A{r}"] = "留本基金总额"
    ws[f"C{r}"] = total_cost + 200_000
    r += 2

    # Bottom summary block (parser optional path)
    summary_header_row = r
    ws[f"A{summary_header_row}"] = "类别"
    ws[f"B{summary_header_row}"] = "投资成本"
    ws[f"C{summary_header_row}"] = "市值"
    ws[f"D{summary_header_row}"] = "增长率"
    r += 1

    def _bucket_totals(col_key: str) -> tuple[float, float]:
        cost = 0.0
        mv = 0.0
        for f, b, _ in FUNDS:
            if b != col_key or f not in lots_by_fund:
                continue
            for sub_date, c, nav, _u in lots_by_fund[f]:
                cost += c
                mv += c * nav
        return cost, mv

    for label, key in [
        ("投资增长类\n(20%)",       "bucket_growth"),
        ("投资固定收益类\n(20%)",   "bucket_fi"),
        ("投资通胀敏感类\n(20%)",   "bucket_inflation"),
        ("投资分散投资类\n(40%)",   "bucket_diversified"),
    ]:
        cost, mv = _bucket_totals(key)
        ws[f"A{r}"] = label
        ws[f"B{r}"] = cost
        ws[f"C{r}"] = mv
        ws[f"D{r}"] = (mv / cost - 1) if cost > 0 else 0
        r += 1


def main() -> None:
    wb: Workbook = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("20260331")
    _write_snapshot(ws1, date(2026, 3, 31), LOTS_20260331)

    ws2 = wb.create_sheet("20260430")
    _write_snapshot(ws2, date(2026, 4, 30), LOTS_20260430)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
