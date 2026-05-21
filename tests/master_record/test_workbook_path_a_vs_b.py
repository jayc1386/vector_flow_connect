"""End-to-end test for `workbook.extract()` Path A vs Path B selection.

Uses the synthetic fixture at
`tests/fixtures/master_record/synthetic_workbook.xlsx` (lifted from
dkup commit `fd67aa8` by plan 0039). The fixture has 2 snapshot
sheets and no `事件流水` sheet — so out-of-the-box it exercises Path
B. The Path A scenario is exercised by programmatically injecting an
events_log sheet into a copy of the fixture before running extract.
"""

from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from vector_flow_connect.master_record import extract
from vector_flow_connect.master_record.events_sheet import DEFAULT_SHEET_NAME

FIXTURE = Path(__file__).parent.parent / "fixtures" / "master_record" / "synthetic_workbook.xlsx"


@pytest.fixture
def synthetic_fixture() -> Path:
    if not FIXTURE.exists():
        pytest.skip(f"synthetic fixture missing: {FIXTURE}")
    return FIXTURE


def test_path_b_when_no_events_log_sheet(synthetic_fixture, tmp_path):
    """The synthetic fixture has no `事件流水` sheet — Path B is the default."""
    result = extract(synthetic_fixture, out_dir=tmp_path)
    manifest = json.loads((tmp_path / "extraction_manifest.json").read_text())

    assert manifest["events_path"] == "B"
    assert manifest["events_log_count"] == 0

    # The synthetic fixture is known to produce 12 subscriptions + 1 redemption
    # (full-redemption-via-notes of 禾禧五号 on 2026-04-18).
    counts = manifest["counts"]["events_by_type"]
    assert counts.get("subscription", 0) == 12
    assert counts.get("redemption", 0) == 1

    # Non-subscription events should carry the explicit lineage flag.
    events = result["events"]
    non_sub = events[events["event_type"] != "subscription"]
    assert (non_sub["data_quality_flag"] == "derived_from_notes").all()

    # Subscription events come from snapshot lot rows, not notes — stay clean.
    sub = events[events["event_type"] == "subscription"]
    assert (sub["data_quality_flag"] == "clean").all()


def test_path_a_when_events_log_present(synthetic_fixture, tmp_path):
    """Inject a 事件流水 sheet into a copy of the fixture and run extract.

    Path A makes the events-log the authoritative source for non-
    subscription events; notes-derived emissions are dropped.
    """
    workbook_copy = tmp_path / "synthetic_with_events_log.xlsx"
    shutil.copyfile(synthetic_fixture, workbook_copy)

    wb = openpyxl.load_workbook(workbook_copy)
    ws = wb.create_sheet(title=DEFAULT_SHEET_NAME)
    ws.append(["日期", "基金代码", "事件类型", "份额变化", "现金流入/出", "单位净值", "备注"])
    # Two events: one redemption + one dividend. The original fixture has
    # one notes-derived redemption; with Path A active, the events-log
    # rows replace it entirely.
    ws.append(
        [date(2026, 4, 18), "000015", "赎回", -1000, 1034.5, 1.0345, "events_log replacement"]
    )
    ws.append([date(2026, 4, 30), "519050", "分红现金", 0, 50, 1.08, "cash distribution"])
    wb.save(workbook_copy)

    result = extract(workbook_copy, out_dir=tmp_path)
    manifest = json.loads((tmp_path / "extraction_manifest.json").read_text())

    assert manifest["events_path"] == "A"
    assert manifest["events_log_count"] == 2

    events = result["events"]
    # Path A: subscriptions still emitted from snapshot (lot identity).
    # Non-subscription events come from the events-log.
    sub = events[events["event_type"] == "subscription"]
    non_sub = events[events["event_type"] != "subscription"]

    # 12 subscriptions from snapshot, unchanged.
    assert len(sub) == 12
    # 2 events-log events: 1 redemption + 1 dividend.
    assert len(non_sub) == 2
    assert (non_sub["data_quality_flag"] == "derived_from_events_log").all()

    # Verify no notes-derived events leaked through.
    assert not (non_sub["data_quality_flag"] == "derived_from_notes").any()


def test_path_a_dedupes_against_existing_subscription_via_event_id(synthetic_fixture, tmp_path):
    """Path A: events_log subscription with same dedup_key as a snapshot
    lot subscription should dedupe (same event_id from same
    (artifact_hash, event_type, dedup_key) tuple).

    The dedup_key shapes are different by design — snapshot uses lot_id,
    events_sheet uses (fund_code, event_date, units_delta). So they
    won't naturally collide. This test documents that behavior: in v1,
    snapshot + events_log subscriptions are independent rows even for
    the same underlying lot, because the dedup_keys are structurally
    different. (Reconciliation between them is a downstream concern,
    not the parser's job.)
    """
    workbook_copy = tmp_path / "synthetic_with_dup_sub.xlsx"
    shutil.copyfile(synthetic_fixture, workbook_copy)

    wb = openpyxl.load_workbook(workbook_copy)
    ws = wb.create_sheet(title=DEFAULT_SHEET_NAME)
    ws.append(["日期", "基金代码", "事件类型", "份额变化", "现金流入/出", "单位净值", "备注"])
    # A 申购 event for an already-known fund. Not expected to dedupe
    # against snapshot-emitted subscriptions in v1.
    ws.append([date(2025, 3, 31), "000015", "申购", 100, -150, 1.5, "duplicate from events_log"])
    wb.save(workbook_copy)

    result = extract(workbook_copy, out_dir=tmp_path)
    manifest = json.loads((tmp_path / "extraction_manifest.json").read_text())

    assert manifest["events_path"] == "A"
    # 12 snapshot subscriptions + 1 events_log subscription = 13
    sub = result["events"][result["events"]["event_type"] == "subscription"]
    assert len(sub) == 13


def test_classify_sheet_recognizes_events_log(synthetic_fixture, tmp_path):
    """The 事件流水 sheet, when present, must classify as 'events_log'
    rather than 'unknown' so it doesn't emit a parser-missing issue.
    """
    workbook_copy = tmp_path / "with_log.xlsx"
    shutil.copyfile(synthetic_fixture, workbook_copy)
    wb = openpyxl.load_workbook(workbook_copy)
    wb.create_sheet(title=DEFAULT_SHEET_NAME)
    wb[DEFAULT_SHEET_NAME].append(
        ["日期", "基金代码", "事件类型", "份额变化", "现金流入/出", "单位净值", "备注"]
    )
    wb[DEFAULT_SHEET_NAME].append([date(2026, 1, 31), "000015", "申购", 100, -150, 1.5, ""])
    wb.save(workbook_copy)

    result = extract(workbook_copy, out_dir=tmp_path)
    manifest = result["manifest"]
    assert manifest["sheet_classification"][DEFAULT_SHEET_NAME] == "events_log"
    # No "unknown_sheet" issue should be emitted for it.
    assert not any(i.get("sheet") == DEFAULT_SHEET_NAME for i in result["issues"])
