"""returns_calc ledger parser — synthetic-workbook tests (no client data)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from vector_flow_connect.dku.returns_calc import LEDGER_SHEET_NAME, extract, parse_ledger

_HEADERS = [
    "标记",
    "日期",
    "借方",
    "已实现",
    "账面余额(已实现）",  # fullwidth closing paren, as in the real sheet
    "主动配置金额变动",
    "主动配置金额",
    "非主动配置金额",
    "总资产市值",
    "总份额",
    "净值",
    "内容",
]


def _build_workbook(path: Path, rows: list[list[object]], *, with_broken_sibling: bool = False):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = LEDGER_SHEET_NAME
    ws.append(["收益计算"])  # title row 1
    ws.append([])  # spacer row 2
    ws.append(_HEADERS)  # header row 3
    for row in rows:
        ws.append(row)
    if with_broken_sibling:
        sib = wb.create_sheet("留本基金波动率计算")
        sib.append(["#REF!", "#VALUE!", "junk"])
    wb.save(path)


def _row(
    mark: str,
    when: datetime | None,
    debit: float,
    realized: str | None,
    units: float,
    nav: float,
    content: str,
    mv: float | None = None,
) -> list[object]:
    mv = mv if mv is not None else units * nav
    return [mark, when, debit, realized, 0, 0, 0, 0, mv, units, nav, content]


_SCENARIO = [
    _row("NA", datetime(2020, 1, 1), 0, "Y", 1000.0, 1.0, "期初余额"),
    _row("NA", datetime(2020, 2, 1), 110.0, "Y", 1100.0, 1.1, "202002捐赠收入-某甲"),
    _row("NA", datetime(2020, 3, 1), 50.0, "Y", 1100.0, 1.15, "2020年一季度收入"),
    _row("分红", datetime(2020, 4, 1), 12.0, "Y", 1100.0, 1.16, "某基金分红"),
    _row("202001", datetime(2020, 5, 1), 8.0, "Y", 1100.0, 1.17, "已实现结构性存款月度收益202001"),
    _row("202002", datetime(2020, 6, 1), 5.0, "N", 1100.0, 1.18, "未实现结构性存款月度收益202002"),
    _row("浮动收益", datetime(2020, 7, 1), -20.0, "N", 1100.0, 1.16, "未实现浮动收益"),
    _row("赎回", datetime(2020, 8, 1), 30.0, "Y", 1100.0, 1.17, "某基金赎回"),
    _row("NA", datetime(2020, 9, 1), 200.0, "Y", 1270.94, 1.17, "可投资产授权"),
]


def test_classification_and_units_delta(tmp_path: Path) -> None:
    wb_path = tmp_path / "05.xlsx"
    # An undated divider row mid-sheet + blank tail rows must be skipped
    # without terminating the scan.
    rows = list(_SCENARIO)
    rows.insert(3, ["", None, None, None, None, None, None, None, None, None, None, "—分割行—"])
    rows.append([None] * 12)
    _build_workbook(wb_path, rows)

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    parsed, issues = parse_ledger(wb[LEDGER_SHEET_NAME], scope="ketou")
    wb.close()

    assert issues == []
    kinds = [r["row_kind"] for r in parsed]
    assert kinds == [
        "opening",
        "external_capital",  # 捐赠 (contains 收入 — gift wins)
        "income_lumped",  # 年/季度收入, realized
        "dividend",
        "interest_realized",
        "unrealized_deposit_interest",
        "unrealized_funds_mtm",
        "internal_move",
        "external_capital",  # 授权 under ketou scope
    ]
    deltas = [r["units_delta"] for r in parsed]
    assert deltas[0] is None  # opening row seeds, never a delta
    assert deltas[1] == pytest.approx(100.0)
    assert deltas[8] == pytest.approx(170.94)
    assert all(d == pytest.approx(0.0) for d in deltas[2:8])
    assert parsed[1]["source_locator"] == f"{LEDGER_SHEET_NAME}!A5"


def test_scope_liuben_authorization_is_internal(tmp_path: Path) -> None:
    wb_path = tmp_path / "04.xlsx"
    _build_workbook(wb_path, _SCENARIO)
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    parsed, _ = parse_ledger(wb[LEDGER_SHEET_NAME], scope="liuben")
    wb.close()
    assert parsed[-1]["row_kind"] == "internal_move"


def test_error_string_cell_nulls_value_and_records_issue(tmp_path: Path) -> None:
    wb_path = tmp_path / "05.xlsx"
    rows = list(_SCENARIO[:2])
    broken = _row("NA", datetime(2020, 10, 1), 1.0, "Y", 1100.0, 1.17, "x")
    broken[8] = "#VALUE!"  # 总资产市值
    rows.append(broken)
    _build_workbook(wb_path, rows)

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    parsed, issues = parse_ledger(wb[LEDGER_SHEET_NAME], scope="ketou")
    wb.close()

    assert parsed[-1]["total_mv"] is None
    assert len(issues) == 1
    assert issues[0]["issue"] == "non_numeric_cell"
    assert issues[0]["column"] == "总资产市值"


def test_wrong_sheet_shape_raises(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = LEDGER_SHEET_NAME
    ws.append(["nothing", "here"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)

    loaded = openpyxl.load_workbook(path, read_only=True, data_only=True)
    with pytest.raises(ValueError, match="追溯 ledger"):
        parse_ledger(loaded[LEDGER_SHEET_NAME], scope="ketou")
    loaded.close()


def test_extract_writes_parquet_manifest_and_ignores_broken_siblings(tmp_path: Path) -> None:
    wb_path = tmp_path / "05.xlsx"
    _build_workbook(wb_path, _SCENARIO, with_broken_sibling=True)

    result = extract(wb_path, out_dir=tmp_path / "out", scope="ketou")

    df = pd.read_parquet(tmp_path / "out" / "ledger_series_ketou.parquet")
    assert len(df) == len(_SCENARIO)
    assert list(df["scope"].unique()) == ["ketou"]
    assert df["source_id"].iloc[0] == "dku_returns_calc_v1"
    assert df["source_artifact"].iloc[0] == "05.xlsx"
    # Multi-date preservation: every scenario row survives (last-per-date
    # selection is downstream's job).
    assert result["manifest"]["counts"]["ledger_rows"] == len(_SCENARIO)
    assert result["manifest"]["counts"]["mint_rows"] == 2  # gift + authorization
    assert (tmp_path / "out" / "extraction_manifest_ketou.json").exists()
    assert (tmp_path / "out" / "issues_ketou.json").exists()


def test_missing_ledger_sheet_raises(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "别的"
    path = tmp_path / "no_ledger.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="no '追溯调整至2022' sheet"):
        extract(path, out_dir=tmp_path / "out", scope="liuben")
