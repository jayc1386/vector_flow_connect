"""Regression test: a bare date in the 持有标的 column must not mint a
phantom position.

Real DKU snapshot sheets place a date cell (e.g. A30 = 2026-04-30) as a
section-divider title for the asset-class summary table that follows the
`留本基金总额` grand total. `_as_text(datetime)` is a non-empty string, so
without an explicit guard the parser treated that date as a fund holding
and emitted a phantom position (fund_id keyed on the stringified date,
units=None, mv≈1.0). The synthetic fixture omits this divider row, so this
case needs its own worksheet.
"""

from __future__ import annotations

from datetime import date, datetime

import openpyxl

from vector_flow_connect.dku.master_record.canonical import SourceContext, fund_id_stub
from vector_flow_connect.dku.master_record.snapshot import parse_sheet


def _build_sheet() -> openpyxl.worksheet.worksheet.Worksheet:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # narrow `Worksheet | None` for pyright
    ws.title = "20260430"

    # Title row (so as_of resolves and matches the sheet name).
    ws["A1"] = "2026年4月30日基金会留本基金资产配置情况"

    # Header row.
    ws["A2"] = "持有标的"
    ws["B2"] = "投资时间"
    ws["C2"] = "投资成本"
    ws["D2"] = "占比"
    ws["E2"] = "投资增长类\n(30%)"
    ws["I2"] = "标的价值"
    ws["J2"] = "单位净值"
    ws["K2"] = "份额"
    ws["L2"] = "备注"

    # One real lot.
    ws["A3"] = "睿远基金睿见1号"
    ws["B3"] = date(2025, 2, 6)
    ws["C3"] = 3_000_000
    ws["D3"] = 1.0
    ws["E3"] = 3_835_961.0  # growth bucket mv (non-zero -> asset_class)
    ws["I3"] = 3_835_961.0
    ws["J3"] = 2.7578
    ws["K3"] = 1_390_950.0

    # Lot block terminators.
    ws["A4"] = "主动配置小计"
    ws["C4"] = 3_000_000
    ws["A5"] = "留本基金总额"
    ws["C5"] = 3_200_000
    # row 6 intentionally blank
    # row 7: the section-divider DATE — the phantom trigger.
    ws["A7"] = datetime(2026, 4, 30)

    # Asset-class summary block (must still be parsed after the date skip).
    ws["A8"] = "类别"
    ws["B8"] = "投资成本"
    ws["C8"] = "市值"
    ws["D8"] = "增长率"
    ws["A9"] = "投资增长类\n(20%)"
    ws["B9"] = 3_000_000
    ws["C9"] = 3_835_961.0
    ws["D9"] = 0.1566

    return ws


def _ctx() -> SourceContext:
    return SourceContext(artifact="synthetic.xlsx", artifact_hash="deadbeef")


def test_date_divider_row_does_not_mint_phantom_position():
    ws = _build_sheet()
    out = parse_sheet(ws, sheet_name="20260430", ctx=_ctx(), resolve_fund_id=fund_id_stub)

    positions = out["positions"]
    # The date divider lives at A7 — no position may originate there.
    assert not [p for p in positions if str(p.get("source_locator", "")).endswith(":A7")]

    # And no position may be keyed on the stringified divider date.
    phantom_fid = fund_id_stub(str(datetime(2026, 4, 30)))
    assert not [p for p in positions if p.get("fund_id") == phantom_fid]


def test_real_lot_and_summary_survive_the_skip():
    ws = _build_sheet()
    out = parse_sheet(ws, sheet_name="20260430", ctx=_ctx(), resolve_fund_id=fund_id_stub)

    # The genuine lot is still captured...
    real = [p for p in out["positions"] if p.get("fund_id") == fund_id_stub("睿远基金睿见1号")]
    assert len(real) == 1

    # ...and skipping the date row did NOT break reaching the 类别 summary
    # header below it (the regression we have to avoid).
    assert out["asset_class_summary"], "asset_class_summary should still parse"
