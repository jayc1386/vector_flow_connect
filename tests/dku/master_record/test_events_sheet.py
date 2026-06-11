"""Unit tests for `vector_flow_connect.dku.master_record.events_sheet`.

Builds small in-memory `openpyxl.Workbook` fixtures inline rather than
loading binary `.xlsx` files. Covers the 6-column 事件流水 shape, the
event-type mapping table, cell-coordinate `source_locator`
construction, and the row-level reconciliation gate.
"""

from __future__ import annotations

from datetime import date, datetime, timezone

import openpyxl
import pytest

from vector_flow_connect.dku.master_record.canonical import SourceContext
from vector_flow_connect.dku.master_record.events_sheet import (
    DEFAULT_SHEET_NAME,
    events_sheet_present,
    parse_events_sheet,
)


def _ctx() -> SourceContext:
    return SourceContext(
        artifact="test.xlsx",
        artifact_hash="deadbeef" * 8,
        extracted_at=datetime(2026, 5, 21, tzinfo=timezone.utc),
    )


def _wb_with_events_sheet(rows: list[list]) -> openpyxl.Workbook:
    """Build an in-memory workbook with the canonical 事件流水 header
    plus the given data rows. Each row is a list of cell values in
    column order matching HEADER_PATTERNS.
    """
    wb = openpyxl.Workbook()
    # openpyxl creates a default sheet "Sheet" — rename it for our use
    ws = wb.active
    ws.title = DEFAULT_SHEET_NAME
    ws.append(["日期", "基金代码", "事件类型", "份额变化", "现金流入/出", "单位净值", "备注"])
    for row in rows:
        ws.append(row)
    return wb


# --- events_sheet_present ---


def test_events_sheet_present_true_when_sheet_and_rows_exist():
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "519050", "申购", 1000, -1500, 1.5, ""]])
    assert events_sheet_present(wb)


def test_events_sheet_present_false_when_sheet_absent():
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    assert not events_sheet_present(wb)


def test_events_sheet_present_false_when_only_header():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DEFAULT_SHEET_NAME
    ws.append(["日期", "基金代码", "事件类型", "份额变化", "现金流入/出", "单位净值", "备注"])
    assert not events_sheet_present(wb)


def test_events_sheet_present_tolerates_whitespace():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"  {DEFAULT_SHEET_NAME}  "
    ws.append(["日期", "基金代码", "事件类型", "份额变化", "现金流入/出", "单位净值", "备注"])
    ws.append([date(2026, 1, 31), "X", "申购", 1, -1, 1, ""])
    assert events_sheet_present(wb)


# --- parse_events_sheet ---


def test_parse_returns_empty_when_sheet_absent():
    wb = openpyxl.Workbook()
    wb.active.title = "20260331"
    assert parse_events_sheet(wb, ctx=_ctx()) == []


def test_event_type_mapping_subscription():
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "519050", "申购", 1000, -1500, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert len(events) == 1
    evt = events[0]
    assert evt["event_type"] == "subscription"
    assert evt["payout_form"] is None
    assert evt["fund_code"] == "519050"
    assert evt["units_delta"] == 1000.0
    assert evt["cash_delta"] == -1500.0
    assert evt["per_unit_amount"] == 1.5


def test_event_type_mapping_dividend_cash_vs_drip():
    wb = _wb_with_events_sheet(
        [
            [date(2026, 1, 31), "A", "分红现金", 0, 100, 1.0, "cash dividend"],
            [date(2026, 1, 31), "B", "分红再投", 50, 0, 1.0, "reinvested"],
        ]
    )
    events = parse_events_sheet(wb, ctx=_ctx())
    assert len(events) == 2
    cash, drip = events
    assert cash["event_type"] == "dividend"
    assert cash["payout_form"] == "cash"
    assert drip["event_type"] == "dividend"
    # Path A now honors the canonical PayoutForm enum ("cash" | "reinvested")
    # instead of the old "drip" alias; matches Path B and prism's
    # record_cash_flow mapping (dividend+reinvested → drip downstream).
    assert drip["payout_form"] == "reinvested"


def test_event_type_mapping_redemption_and_perf_fee():
    wb = _wb_with_events_sheet(
        [
            [date(2026, 2, 15), "A", "赎回", -1000, 1500, 1.5, ""],
            [date(2026, 2, 28), "A", "业绩报酬", 0, -50, None, ""],
        ]
    )
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["event_type"] == "redemption"
    assert events[0]["payout_form"] is None
    assert events[1]["event_type"] == "perf_fee"
    assert events[1]["payout_form"] is None
    # perf_fee subtype rides into notes_raw with a structured prefix
    # instead of overloading payout_form.
    assert "[perf_fee_subtype=performance_fee]" in events[1]["notes_raw"]


def test_fee_subtypes_ride_into_notes_raw_not_payout_form():
    wb = _wb_with_events_sheet(
        [
            [date(2026, 3, 1), "A", "申购费", 0, -50, None, ""],
            [date(2026, 3, 1), "A", "赎回费", 0, -30, None, ""],
        ]
    )
    events = parse_events_sheet(wb, ctx=_ctx())
    for evt in events:
        # payout_form is reserved for dividend events; perf_fee carries None
        # and the entry/exit-fee subtype goes to notes_raw.
        assert evt["event_type"] == "perf_fee"
        assert evt["payout_form"] is None
    assert "[perf_fee_subtype=entry_fee]" in events[0]["notes_raw"]
    assert "[perf_fee_subtype=exit_fee]" in events[1]["notes_raw"]


def test_unknown_event_type_emits_flagged_row_not_dropped():
    wb = _wb_with_events_sheet([[date(2026, 3, 1), "X", "未知类型", 0, 0, None, ""]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert len(events) == 1
    # Falls back to perf_fee; unknown label rides into notes_raw under the
    # perf_fee_subtype prefix (payout_form stays None to keep the dividend-
    # only enum clean).
    assert events[0]["event_type"] == "perf_fee"
    assert events[0]["payout_form"] is None
    assert "unknown_event_type:未知类型" in events[0]["notes_raw"]


# --- source_locator construction ---


def test_source_locator_includes_sheet_and_cell_coord():
    wb = _wb_with_events_sheet(
        [
            [date(2026, 1, 31), "519050", "申购", 1000, -1500, 1.5, ""],
            [date(2026, 2, 15), "519050", "赎回", -1000, 1500, 1.5, ""],
        ]
    )
    events = parse_events_sheet(wb, ctx=_ctx())
    # Header is row 1, data rows are 2 and 3. 日期 is column A.
    assert events[0]["source_locator"] == f"{DEFAULT_SHEET_NAME}!A2"
    assert events[1]["source_locator"] == f"{DEFAULT_SHEET_NAME}!A3"


# --- reconciliation gate ---


def test_reconcile_clean_when_invariant_holds():
    # 1000 units * 1.5 NAV = 1500 cash (sign-insensitive abs)
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "A", "申购", 1000, -1500, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["data_quality_flag"] == "derived_from_events_log"


def test_reconcile_flags_cash_share_mismatch_when_invariant_fails():
    # 1000 * 1.5 = 1500, but cash recorded as -2000 (33% off — well past 1%)
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "A", "申购", 1000, -2000, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["data_quality_flag"] == "cash_share_mismatch"


def test_reconcile_within_1pct_tolerance_stays_clean():
    # 1000 * 1.5 = 1500, cash recorded as -1505 (0.33% off — within tolerance)
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "A", "申购", 1000, -1505, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["data_quality_flag"] == "derived_from_events_log"


def test_reconcile_inapplicable_when_field_missing():
    # cash_delta zero → gate inapplicable, row treated as clean
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "A", "分红再投", 50, 0, 1.0, "drip"]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["data_quality_flag"] == "derived_from_events_log"


# --- provenance stamping ---


def test_provenance_fields_stamped_from_context():
    ctx = _ctx()
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "A", "申购", 1000, -1500, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=ctx)
    evt = events[0]
    assert evt["source_artifact"] == ctx.artifact
    assert evt["source_artifact_hash"] == ctx.artifact_hash
    assert evt["extracted_at"] == ctx.extracted_at
    # recorded_at is day-grain to match snapshot.py's convention
    # (cross-source events must be sortable in one DataFrame column).
    assert evt["recorded_at"] == ctx.extracted_at.date()
    assert evt["source_id"] == "dku_master_record_v1"
    assert evt["schema_version"] == "dku-master-record-v1"
    assert evt["extractor_name"] == "dku_master_record"
    assert evt["currency"] == "CNY"


def test_event_id_stable_across_reads():
    ctx = _ctx()
    wb = _wb_with_events_sheet([[date(2026, 1, 31), "A", "申购", 1000, -1500, 1.5, ""]])
    e1 = parse_events_sheet(wb, ctx=ctx)
    e2 = parse_events_sheet(wb, ctx=ctx)
    assert e1[0]["event_id"] == e2[0]["event_id"]


def test_event_id_changes_when_units_or_date_change():
    ctx = _ctx()
    wb_a = _wb_with_events_sheet([[date(2026, 1, 31), "A", "申购", 1000, -1500, 1.5, ""]])
    wb_b = _wb_with_events_sheet([[date(2026, 2, 1), "A", "申购", 1000, -1500, 1.5, ""]])
    a = parse_events_sheet(wb_a, ctx=ctx)
    b = parse_events_sheet(wb_b, ctx=ctx)
    assert a[0]["event_id"] != b[0]["event_id"]


# --- coercion ---


def test_string_decimals_parsed_with_commas_and_signs():
    wb = _wb_with_events_sheet(
        [[date(2026, 1, 31), "A", "申购", "+1,234.56", "-1,851.84", "1.5", ""]]
    )
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["units_delta"] == pytest.approx(1234.56)
    assert events[0]["cash_delta"] == pytest.approx(-1851.84)


def test_iso_string_date_parsed():
    wb = _wb_with_events_sheet([["2026-01-31", "A", "申购", 1000, -1500, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=_ctx())
    assert events[0]["event_date"] == date(2026, 1, 31)


def test_blank_rows_skipped():
    wb = _wb_with_events_sheet(
        [
            [date(2026, 1, 31), "A", "申购", 1000, -1500, 1.5, ""],
            [None, None, None, None, None, None, None],
            [date(2026, 2, 1), "A", "赎回", -1000, 1500, 1.5, ""],
        ]
    )
    events = parse_events_sheet(wb, ctx=_ctx())
    assert len(events) == 2


def test_custom_resolver_passed_through():
    resolved: list[str] = []

    def fake_resolver(code: str) -> str:
        resolved.append(code)
        return f"fnd_test_{code}"

    wb = _wb_with_events_sheet([[date(2026, 1, 31), "519050", "申购", 1000, -1500, 1.5, ""]])
    events = parse_events_sheet(wb, ctx=_ctx(), resolve_fund_id=fake_resolver)
    assert resolved == ["519050"]
    assert events[0]["fund_id"] == "fnd_test_519050"
