"""Parser for the proposed 事件流水 (events log) sheet.

Per DKU_QUESTIONS.md ask #5, DKU is asked to record portfolio events
as one row per event in a dedicated sheet (rather than as cumulative
totals in the existing 注释/备注 free-text columns). The proposed
shape:

    | 日期 | 基金代码 | 事件类型 | 份额变化 | 现金流入/出 | 单位净值 | 备注 |
    | 2026-01-31 | SGV901 | 申购 | +12,345.67 | -15,000.00 | 1.2151 |  |

This parser turns those rows into canonical events with explicit
provenance (`source_locator = "事件流水!{cell}"`) and the
`data_quality_flag = "derived_from_events_log"` lineage marker.

**v1 posture: speculative.** DKU's earliest realistic adoption is
~2026-06; today's workbooks omit this sheet entirely.
`events_sheet_present()` is the feature-detection switch that
lets `workbook.extract()` activate Path A automatically when the
sheet appears.

Pure-Python; no I/O beyond reading openpyxl cells.
"""

from __future__ import annotations

from collections.abc import Callable
from datetime import date, datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ._inherited_canonical_contract import validate_data_quality_flag
from .canonical import (
    EXTRACTOR_NAME,
    EXTRACTOR_VERSION,
    SCHEMA_VERSION,
    SOURCE_ID,
    SourceContext,
    empty_event,
    event_id,
    fund_id_stub,
)

DEFAULT_SHEET_NAME = "事件流水"

# Column-header → canonical-event-field mapping.
# Header text is matched via substring (case- and whitespace-tolerant)
# so DKU can vary the header lightly without breaking the parser.
HEADER_PATTERNS: dict[str, str] = {
    "日期": "event_date",
    "基金代码": "fund_code",
    "事件类型": "event_type_zh",
    "份额变化": "units_delta",
    "现金流入": "cash_delta",  # matches both "现金流入/出" and "现金流入出"
    "单位净值": "per_unit_amount",
    "备注": "notes_raw",
}

# Chinese event-type label → canonical (event_type, payout_form) tuple.
# canonical event_type ∈ {"subscription", "redemption", "dividend", "perf_fee"}.
EVENT_TYPE_MAP: dict[str, tuple[str, str | None]] = {
    "申购": ("subscription", None),
    "赎回": ("redemption", None),
    "分红现金": ("dividend", "cash"),
    "分红再投": ("dividend", "drip"),
    "业绩报酬": ("perf_fee", None),
    "申购费": ("perf_fee", "entry_fee"),
    "赎回费": ("perf_fee", "exit_fee"),
}

# Reconciliation gate tolerance for `|份额变化 × 单位净值| ≈ |现金流入/出|`.
# 1% absolute tolerance on the cash side. Below this, the row is clean;
# above, the row's `data_quality_flag` flips to `cash_share_mismatch`.
RECON_TOLERANCE = 0.01


def events_sheet_present(wb: openpyxl.Workbook, *, sheet_name: str = DEFAULT_SHEET_NAME) -> bool:
    """Feature-detect the optional 事件流水 sheet.

    Returns True if a sheet with exactly `sheet_name` exists in the
    workbook and is non-empty (has at least one row after the
    header). Whitespace in the title is normalized; case is not.
    """
    target = sheet_name.strip()
    for actual in wb.sheetnames:
        if actual.strip() == target:
            ws = wb[actual]
            return bool(ws.max_row and ws.max_row > 1)
    return False


def parse_events_sheet(
    wb: openpyxl.Workbook,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
    ctx: SourceContext,
    resolve_fund_id: Callable[[str], str] | None = None,
) -> list[dict]:
    """Parse the events-log sheet into a list of canonical event dicts.

    Each returned dict matches `EVENT_COLUMNS` shape with:
      - `source_locator = "事件流水!{cell}"` pointing at the 日期 cell
        of the originating row.
      - `data_quality_flag = "derived_from_events_log"` (escalated to
        `"cash_share_mismatch"` when the row-level invariant fails).
      - `event_id` deterministic from
        `(artifact_hash, event_type, fund_code, event_date, units_delta)`
        so re-extracts of the same row dedupe natively.
      - All standard provenance fields stamped from `ctx`.

    `resolve_fund_id` resolves a fund_code string to a stable
    canonical `fund_id`. If `None`, falls back to
    `fund_id_stub(fund_code or "")`. The same resolver shape as
    `snapshot.parse_sheet`.
    """
    if not events_sheet_present(wb, sheet_name=sheet_name):
        return []

    ws: Worksheet = wb[sheet_name]
    header_to_col = _map_headers(ws)
    if not header_to_col:
        return []

    resolver = resolve_fund_id or (lambda s: fund_id_stub(s or ""))

    events: list[dict] = []
    for row_idx in range(2, (ws.max_row or 1) + 1):
        row_dict = _extract_row(ws, row_idx, header_to_col)
        if row_dict is None:
            continue  # blank row; skip silently

        event_type_zh = row_dict.get("event_type_zh", "").strip()
        canon = EVENT_TYPE_MAP.get(event_type_zh)
        if canon is None:
            # Unknown event-type — emit a flagged row so the operator
            # sees it instead of silently dropping it.
            canon = ("perf_fee", f"unknown_event_type:{event_type_zh}")

        event_type, payout_form_or_qualifier = canon

        # Reconciliation gate (sign-insensitive). Only meaningful for
        # transactions where both units AND cash move (subscription,
        # redemption); cash dividends + perf fees have units_delta=0 by
        # construction and shouldn't be gated against the invariant.
        if event_type in ("subscription", "redemption"):
            dq_flag = _reconcile_row(
                units_delta=row_dict.get("units_delta"),
                cash_delta=row_dict.get("cash_delta"),
                per_unit_amount=row_dict.get("per_unit_amount"),
            )
        else:
            dq_flag = "derived_from_events_log"

        # Source locator: the 日期 column's cell coordinate (e.g. "A14"),
        # prefixed with the sheet name.
        date_col = header_to_col.get("event_date")
        if date_col is not None:
            cell_coord = ws.cell(row=row_idx, column=date_col).coordinate
            source_locator = f"{ws.title}!{cell_coord}"
        else:
            source_locator = f"{ws.title}!row={row_idx}"

        fund_code = row_dict.get("fund_code") or None
        fund_id = resolver(fund_code) if fund_code else resolver(event_type_zh)

        # Deterministic event_id — re-extracts of the same row produce
        # the same id. Uses fund_code (or "" if absent), the canonical
        # event_type, the event date, and the units_delta to discriminate.
        dedup_key = "::".join(
            [
                fund_code or "",
                str(row_dict.get("event_date") or ""),
                str(row_dict.get("units_delta") or ""),
            ]
        )

        evt = empty_event()
        evt.update(
            event_id=event_id(ctx.artifact_hash, event_type, dedup_key),
            event_type=event_type,
            fund_id=fund_id,
            fund_code=fund_code,
            source_fund_string=None,  # not in events-log shape; only fund_code
            event_date=row_dict.get("event_date"),
            valid_from=row_dict.get("event_date"),
            units_delta=row_dict.get("units_delta"),
            cash_delta=row_dict.get("cash_delta"),
            per_unit_amount=row_dict.get("per_unit_amount"),
            payout_form=payout_form_or_qualifier,
            currency="CNY",
            confidence="clean",
            data_quality_flag=validate_data_quality_flag(dq_flag),
            notes_raw=row_dict.get("notes_raw"),
            source_artifact=ctx.artifact,
            source_artifact_hash=ctx.artifact_hash,
            source_locator=source_locator,
            source_id=SOURCE_ID,
            extractor_name=EXTRACTOR_NAME,
            extractor_version=EXTRACTOR_VERSION,
            schema_version=SCHEMA_VERSION,
            extracted_at=ctx.extracted_at,
            # snapshot.py uses a `date` for `recorded_at` (the sheet's
            # as_of date); match that grain here so cross-source events
            # are sortable in the same column.
            recorded_at=ctx.extracted_at.date(),
        )
        events.append(evt)

    return events


# --- internals ---


def _map_headers(ws: Worksheet) -> dict[str, int]:
    """Match row-1 header cells against HEADER_PATTERNS via substring.

    Returns a mapping from canonical field name → 1-based column index.
    Fields not found in the header row are absent from the returned dict.
    """
    out: dict[str, int] = {}
    if ws.max_column is None or ws.max_column == 0:
        return out
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col_idx).value
        if not isinstance(raw, str):
            continue
        cell_text = raw.strip()
        for pattern, field in HEADER_PATTERNS.items():
            if pattern in cell_text and field not in out:
                out[field] = col_idx
                break
    return out


def _extract_row(ws: Worksheet, row_idx: int, header_to_col: dict[str, int]) -> dict | None:
    """Read one event row, return field-dict or None if blank.

    Coerces:
      - 日期 → date (handles datetime, date, ISO string)
      - 基金代码 → stripped str (or None)
      - 事件类型 → stripped str
      - 份额变化 / 现金流入/出 / 单位净值 → Decimal | None
      - 备注 → str | None
    """
    fields: dict = {}
    has_any = False

    for field, col in header_to_col.items():
        raw = ws.cell(row=row_idx, column=col).value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            fields[field] = None
            continue

        has_any = True

        if field == "event_date":
            fields[field] = _coerce_date(raw)
        elif field in ("units_delta", "cash_delta", "per_unit_amount"):
            fields[field] = _coerce_float(raw)
        elif field == "fund_code":
            fields[field] = str(raw).strip()
        elif field in ("event_type_zh", "notes_raw"):
            fields[field] = str(raw).strip() if isinstance(raw, str) else str(raw)
        else:
            fields[field] = raw

    return fields if has_any else None


def _coerce_date(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        s = raw.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _coerce_float(raw: object) -> float | None:
    """Coerce a workbook cell value to a Python float.

    Aligns with snapshot.py's float-based parquet shape so events from
    both sources can co-exist in one DataFrame without dtype churn.
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        s = raw.strip().replace(",", "").replace("+", "")
        if not s:
            return None
        try:
            return float(s)
        except (ValueError, ArithmeticError):
            return None
    return None


def _reconcile_row(
    *,
    units_delta: float | None,
    cash_delta: float | None,
    per_unit_amount: float | None,
) -> str:
    """Apply the row-level invariant `|units × nav| ≈ |cash|`.

    Returns `"derived_from_events_log"` when the gate passes (clean
    lineage) or `"cash_share_mismatch"` when the gate fails. If any
    of the three inputs is missing or cash_delta is zero, the gate is
    inapplicable and the row is treated as clean-by-default.
    """
    if units_delta is None or cash_delta is None or per_unit_amount is None or cash_delta == 0:
        return "derived_from_events_log"
    try:
        expected = abs(units_delta * per_unit_amount)
        observed = abs(cash_delta)
        diff = abs(expected - observed)
        tolerance = RECON_TOLERANCE * observed
        if diff > tolerance:
            return "cash_share_mismatch"
        return "derived_from_events_log"
    except (ValueError, ArithmeticError):
        return "cash_share_mismatch"
