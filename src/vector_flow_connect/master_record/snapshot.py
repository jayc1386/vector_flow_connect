"""Parser for YYYYMMDD snapshot sheets in 留本基金动态资产配置情况.xlsx.

Each snapshot sheet records the endowment's position lots as of a single
date. The parser emits:

- one **subscription event** per lot row (deduplicated across snapshots
  in the workbook-level dispatcher);
- one **position observation** per lot row (separate output table,
  feeds reconciliation);
- zero-or-more **notes events** per lot row via `notes_parser` —
  redemption / dividend / DRIP / perf-fee.

Column schema is **heterogeneous across eras** — early 2020 sheets have
~10 columns; 2024+ sheets have ~18. Columns are resolved by partial
substring match on the header strings; missing columns silently
produce None values in the output.
"""

from __future__ import annotations

import re
from collections.abc import Callable
from datetime import date, datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from . import notes_parser
from ._inherited_canonical_contract import validate_asset_class
from .canonical import (
    EXTRACTOR_NAME,
    EXTRACTOR_VERSION,
    SCHEMA_VERSION,
    SOURCE_ID,
    SourceContext,
    empty_event,
    empty_position,
    event_id,
    fund_id_stub,
    lot_id,
)

# Substring → canonical-column-name mapping. Match order matters where
# a label might match more than one entry (e.g. 投资固定收益类 also
# contains 投资 — list more-specific patterns first).
COL_PATTERNS: list[tuple[str, str]] = [
    ("持有标的", "fund_string"),
    ("投资时间", "subscription_date"),
    ("投资成本", "cost"),
    ("占比", "weight_within_fund"),
    ("投资增长类", "bucket_growth"),
    ("投资固定收益类", "bucket_fi"),
    ("投资债券与现金类", "bucket_fi"),  # earlier taxonomy alias
    ("投资通胀敏感类", "bucket_inflation"),
    ("投资分散投资类", "bucket_diversified"),
    ("单位净值", "nav"),
    ("份额", "units"),
    ("标的价值", "mv"),
    ("标的增长额", "pnl"),
    ("标的年化回报率", "ann_return"),
    ("标的回报", "ann_return"),  # early-era label
    ("比列", "position_weight"),
    ("持有天数", "holding_days"),
    ("注释", "notes"),
    ("备注", "notes"),
    ("总结", "summary"),
]

# First-column markers that end the lot section.
SUBTOTAL_MARKERS: set[str] = {
    "主动配置小计",
    "现金产品小计",
    "留本基金总额",
    "总额",  # earliest sheets
    "配置比例",  # asset-class summary header in early sheets
}

# First-column markers that *aren't* lots but appear in the lot region
# of early sheets (uninvested cash). Treated as position observations
# but NOT emitted as subscription events.
NON_LOT_MARKERS: set[str] = {
    "未投现金",
}

# Asset-class block headers — recognized so we can parse the bottom
# summary block.
SUMMARY_HEADER_FIRST_COL = {"类别"}

TITLE_RE = re.compile(r"^\s*(\d{4})年(\d{1,2})月(\d{1,2})日.*?基金会留本基金")


def _as_date(value: Any) -> date | None:
    """Coerce openpyxl date-ish values to a `date`."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _as_float(value: Any) -> float | None:
    if value is None or value == "" or value == "-":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _bucket_from_buckets(row: dict[str, Any]) -> str | None:
    """The non-zero bucket column tells us the asset class."""
    mapping = {
        "bucket_growth": "Growth",
        "bucket_fi": "Fixed Income",
        "bucket_inflation": "Inflation Sensitive",
        "bucket_diversified": "Diversifiers",
    }
    for col, label in mapping.items():
        val = row.get(col)
        if val not in (None, 0, 0.0):
            return validate_asset_class(label)
    return None


def _find_title_date(ws: Worksheet) -> date | None:
    """Look in the first few rows for the title cell with a parseable date."""
    for row in ws.iter_rows(min_row=1, max_row=4, values_only=False):
        for cell in row:
            if isinstance(cell.value, str):
                m = TITLE_RE.match(cell.value)
                if m:
                    y, mth, d = (int(x) for x in m.groups())
                    return date(y, mth, d)
    return None


def _find_header_row(ws: Worksheet) -> tuple[int, dict[int, str]]:
    """Find the row that contains `持有标的` and return (row_idx, col_idx → canonical_name).

    Returns (-1, {}) if not found — caller treats as fatal.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=False), start=1):
        first = row[0].value if row[0].value else ""
        if isinstance(first, str) and "持有标的" in first:
            col_map: dict[int, str] = {}
            for cell in row:
                col_idx = getattr(cell, "column", None)
                if col_idx is None or not isinstance(cell.value, str):
                    continue
                label = cell.value.replace("\n", "").strip()
                # match against COL_PATTERNS — first match wins
                for substr, canon in COL_PATTERNS:
                    if substr in label:
                        col_map[col_idx] = canon
                        break
            return row_idx, col_map
    return -1, {}


def scan_fund_strings(ws: Worksheet) -> list[str]:
    """Fast pass — return every distinct fund_string seen in the sheet's
    `持有标的` column. Used by `workbook.extract` to build the fund
    identity resolver before the main parse pass.
    """
    header_row_idx, col_map = _find_header_row(ws)
    if header_row_idx < 0:
        return []
    # find the column holding fund_string
    fund_col_idx = None
    for col_idx, canon in col_map.items():
        if canon == "fund_string":
            fund_col_idx = col_idx
            break
    if fund_col_idx is None:
        return []
    out: list[str] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if getattr(cell, "column", None) != fund_col_idx:
                continue
            val = cell.value
            if not isinstance(val, str):
                continue
            v = val.replace("\n", " ").strip()
            if not v:
                continue
            # stop at first subtotal marker — anything past it is summary
            if v in SUBTOTAL_MARKERS or v in SUMMARY_HEADER_FIRST_COL:
                return out
            out.append(v)
            break
    return out


def parse_sheet(
    ws: Worksheet,
    *,
    sheet_name: str,
    ctx: SourceContext,
    resolve_fund_id: Callable[[str], str] = fund_id_stub,
) -> dict[str, list[dict]]:
    """Parse one snapshot sheet.

    Returns:
        {
            "events": [subscription events + notes-parser events ...],
            "positions": [per-lot position observations ...],
            "asset_class_summary": [{asset_class, cost, mv, growth_rate, target_weight} ...],
            "issues": [{sheet, locator, kind, detail} ...],
        }
    """
    issues: list[dict] = []
    events: list[dict] = []
    positions: list[dict] = []
    observations: list[dict] = []
    asset_summary: list[dict] = []

    title_date = _find_title_date(ws)
    # cross-check against sheet name
    sheet_date: date | None = None
    if re.fullmatch(r"\d{8}", sheet_name):
        sheet_date = date(int(sheet_name[:4]), int(sheet_name[4:6]), int(sheet_name[6:8]))
    as_of = title_date or sheet_date
    if title_date and sheet_date and title_date != sheet_date:
        issues.append(
            {
                "sheet": sheet_name,
                "locator": f"{sheet_name}:title",
                "kind": "title_date_mismatch",
                "detail": f"title={title_date} sheet_name={sheet_date}",
            }
        )
    if as_of is None:
        issues.append(
            {
                "sheet": sheet_name,
                "locator": f"{sheet_name}:title",
                "kind": "no_as_of_date",
                "detail": "neither title nor sheet name parsed",
            }
        )
        return {"events": [], "positions": [], "asset_class_summary": [], "issues": issues}

    header_row_idx, col_map = _find_header_row(ws)
    if header_row_idx < 0:
        issues.append(
            {
                "sheet": sheet_name,
                "locator": f"{sheet_name}:header",
                "kind": "no_header_row",
                "detail": "could not find row containing 持有标的",
            }
        )
        return {"events": [], "positions": [], "asset_class_summary": [], "issues": issues}

    # Walk lot rows.
    summary_header_row_idx: int | None = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=False),
        start=header_row_idx + 1,
    ):
        first_val = row[0].value if row else None
        first_text = _as_text(first_val)

        # Stop conditions.
        if first_text is None or first_text.strip() == "":
            # Skip blank rows but don't stop — early-era sheets have a
            # blank between lots and 总额 sometimes.
            continue
        first_clean = first_text.strip()
        if first_clean in SUMMARY_HEADER_FIRST_COL:
            summary_header_row_idx = row_idx
            break
        if first_clean in SUBTOTAL_MARKERS:
            # capture subtotal as an issue marker (useful for reconcile)
            # but skip event emission
            continue

        # Build a column-keyed dict of cell values for this row.
        row_dict: dict[str, Any] = {}
        coord_dict: dict[str, str] = {}
        for cell in row:
            col_idx = getattr(cell, "column", None)
            if col_idx is None:
                continue
            canon = col_map.get(col_idx)
            if canon:
                row_dict[canon] = cell.value
                coord_dict[canon] = getattr(cell, "coordinate", f"col{col_idx}{row_idx}")

        fund_string = _as_text(row_dict.get("fund_string"))
        if not fund_string:
            continue
        # Strip embedded newlines from the fund cell (e.g. "华夏纯债债券A\n(000015)")
        fund_string_clean = fund_string.replace("\n", " ").strip()
        # parse out (XXXXXX) fund code if present.
        # NOTE: fund_code and name_zh are parsed here for future use
        # (and for the funds-registry build pass in `workbook.py`) but
        # not consumed in this function itself.
        code_match = re.search(r"\((\d{4,6})\)", fund_string_clean)
        fund_code = code_match.group(1) if code_match else None  # noqa: F841
        name_zh = re.sub(r"\(\d{4,6}\)", "", fund_string_clean).strip()  # noqa: F841

        # Treat non-lot rows (现金 etc.) as positions only.
        is_lot = first_clean not in NON_LOT_MARKERS

        sub_date = _as_date(row_dict.get("subscription_date"))
        cost = _as_float(row_dict.get("cost"))
        nav = _as_float(row_dict.get("nav"))
        units = _as_float(row_dict.get("units"))
        mv = _as_float(row_dict.get("mv"))
        pnl = _as_float(row_dict.get("pnl"))
        ann_return = _as_float(row_dict.get("ann_return"))
        weight_within_fund = _as_float(row_dict.get("weight_within_fund"))
        position_weight = _as_float(row_dict.get("position_weight"))
        holding_days = _as_float(row_dict.get("holding_days"))
        notes_text = _as_text(row_dict.get("notes"))
        summary_text = _as_text(row_dict.get("summary"))
        asset_class = _bucket_from_buckets(row_dict)

        fund_id = resolve_fund_id(fund_string_clean)

        # Lot identity requires a subscription date + cost.
        this_lot_id: str | None = None
        if is_lot and sub_date is not None and cost is not None:
            this_lot_id = lot_id(fund_id, sub_date, cost)

        # --- Subscription event ---
        if is_lot and this_lot_id is not None:
            sub_locator = f"{sheet_name}:{coord_dict.get('fund_string', f'A{row_idx}')}"
            evt = empty_event()
            evt.update(
                # Subscription event_id is per-LOT, not per-sheet, so the
                # same subscription seen across many snapshots dedupes
                # naturally to a single canonical event row.
                event_id=event_id(ctx.artifact_hash, "subscription", this_lot_id),
                event_type="subscription",
                fund_id=fund_id,
                source_fund_string=fund_string_clean,
                lot_id=this_lot_id,
                event_date=sub_date,
                valid_from=sub_date,
                # Subscription is a -cost cashflow and +units share creation.
                units_delta=units,
                cash_delta=(-cost) if cost is not None else None,
                per_unit_amount=(cost / units) if (cost is not None and units) else nav,
                currency="CNY",
                confidence="clean",
                data_quality_flag="clean",
                source_artifact=ctx.artifact,
                source_artifact_hash=ctx.artifact_hash,
                source_locator=sub_locator,
                source_id=SOURCE_ID,
                extractor_name=EXTRACTOR_NAME,
                extractor_version=EXTRACTOR_VERSION,
                schema_version=SCHEMA_VERSION,
                extracted_at=ctx.extracted_at,
                recorded_at=as_of,
            )
            events.append(evt)

        # --- Position observation ---
        pos = empty_position()
        pos.update(
            lot_id=this_lot_id,
            fund_id=fund_id,
            as_of=as_of,
            units=units,
            nav=nav,
            mv=mv,
            cost=cost,
            pnl=pnl,
            ann_return=ann_return,
            weight_within_fund=weight_within_fund,
            position_weight=position_weight,
            asset_class=asset_class,
            holding_days=holding_days,
            data_quality_flag="clean",
            notes_raw=notes_text,
            source_artifact=ctx.artifact,
            source_artifact_hash=ctx.artifact_hash,
            source_locator=f"{sheet_name}:A{row_idx}",
            source_id=SOURCE_ID,
            extractor_name=EXTRACTOR_NAME,
            extractor_version=EXTRACTOR_VERSION,
            schema_version=SCHEMA_VERSION,
            extracted_at=ctx.extracted_at,
        )
        positions.append(pos)

        # --- Notes parsing on both 注释 and 总结 columns ---
        if this_lot_id is not None:
            for col_name, col_text in (("notes", notes_text), ("summary", summary_text)):
                if not col_text:
                    continue
                loc = f"{sheet_name}:{coord_dict.get(col_name, f'?{row_idx}')}"
                events.extend(
                    notes_parser.parse(
                        col_text,
                        lot_context={
                            "lot_id": this_lot_id,
                            "fund_id": fund_id,
                            "source_fund_string": fund_string_clean,
                            "units_at_lot": units,
                        },
                        source_locator=loc,
                        source_artifact=ctx.artifact,
                        source_artifact_hash=ctx.artifact_hash,
                        recorded_at=as_of,
                        extracted_at=ctx.extracted_at,
                    )
                )
                # Fund-level balance observations (cumulative DRIP, ...)
                observations.extend(
                    notes_parser.parse_observations(
                        col_text,
                        fund_id=fund_id,
                        source_fund_string=fund_string_clean,
                        source_locator=loc,
                        recorded_at=as_of,
                        source_artifact=ctx.artifact,
                        source_artifact_hash=ctx.artifact_hash,
                        extracted_at=ctx.extracted_at,
                    )
                )

    # --- Asset-class summary block ---
    if summary_header_row_idx is not None:
        # Header row layout (20260430): A=类别 B=投资成本 C=市值 D=增长率
        # Rows continue until 合计.
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=summary_header_row_idx + 1, max_row=ws.max_row, values_only=False),
            start=summary_header_row_idx + 1,
        ):
            label = _as_text(row[0].value)
            if not label:
                continue
            label_clean = label.replace("\n", " ").strip()
            cost_val = _as_float(row[1].value if len(row) > 1 else None)
            mv_val = _as_float(row[2].value if len(row) > 2 else None)
            growth_val = _as_float(row[3].value if len(row) > 3 else None)
            # Try to parse the target_weight % from the label, e.g. "(20%)" / "(0-30%)"
            target_weight = None
            target_weight_range = None
            tw_match = re.search(r"\((\d+(?:-\d+)?)%\)", label_clean)
            if tw_match:
                raw = tw_match.group(1)
                if "-" in raw:
                    lo, hi = raw.split("-")
                    target_weight_range = (float(lo) / 100, float(hi) / 100)
                else:
                    target_weight = float(raw) / 100

            asset_summary.append(
                {
                    "as_of": as_of,
                    "label_raw": label_clean,
                    "asset_class": _normalize_asset_class_label(label_clean),
                    "cost": cost_val,
                    "mv": mv_val,
                    "growth_rate": growth_val,
                    "target_weight": target_weight,
                    "target_weight_range": target_weight_range,
                    "source_locator": f"{sheet_name}:A{row_idx}",
                }
            )
            if label_clean.startswith("合计"):
                break

    return {
        "events": events,
        "positions": positions,
        "observations": observations,
        "asset_class_summary": asset_summary,
        "issues": issues,
    }


def _normalize_asset_class_label(label: str) -> str | None:
    if "增长" in label:
        return validate_asset_class("Growth")
    if "固定收益" in label or "债券与现金" in label:
        return validate_asset_class("Fixed Income")
    if "通胀" in label:
        return validate_asset_class("Inflation Sensitive")
    if "分散" in label:
        return validate_asset_class("Diversifiers")
    return None
