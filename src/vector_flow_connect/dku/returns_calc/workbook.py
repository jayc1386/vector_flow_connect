"""Top-level extract() for the 收益计算 workbooks.

Opens ONLY the 追溯 ledger sheet — sibling sheets carry known #REF! /
#VALUE! breakage (波动率 / 市值变化 / 不追溯) and must never reach the
parser. Outputs are suffixed by scope so the 04 and 05 extracts can
share an out_dir.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd

from vector_flow_connect.extraction_contract import file_sha256

from .canonical import (
    EXTRACTOR_NAME,
    EXTRACTOR_VERSION,
    LEDGER_COLUMNS,
    SCHEMA_VERSION,
    SOURCE_ID,
    Scope,
)
from .ledger import LEDGER_SHEET_NAME, parse_ledger


def extract(
    workbook_path: str | Path,
    *,
    out_dir: str | Path,
    scope: Scope,
) -> dict:
    """Parse the ledger and write `ledger_series_{scope}.parquet` +
    `extraction_manifest_{scope}.json` + `issues_{scope}.json`.

    Returns ``{"ledger": DataFrame, "manifest": dict, "issues": list}``.
    """
    workbook_path = Path(workbook_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    artifact_hash = file_sha256(workbook_path)
    extracted_at = datetime.now(timezone.utc)

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if LEDGER_SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"{workbook_path.name} has no {LEDGER_SHEET_NAME!r} sheet (sheets: {wb.sheetnames})"
            )
        rows, issues = parse_ledger(wb[LEDGER_SHEET_NAME], scope=scope)
    finally:
        wb.close()

    for row in rows:
        row["source_artifact"] = workbook_path.name
        row["source_artifact_hash"] = artifact_hash
        row["source_id"] = SOURCE_ID
        row["extractor_name"] = EXTRACTOR_NAME
        row["extractor_version"] = EXTRACTOR_VERSION
        row["schema_version"] = SCHEMA_VERSION
        row["extracted_at"] = extracted_at.isoformat()

    ledger_df = pd.DataFrame(rows, columns=LEDGER_COLUMNS)

    mint_rows = (
        int((ledger_df["units_delta"].fillna(0.0).abs() > 1e-6).sum()) if not ledger_df.empty else 0
    )
    manifest = {
        "workbook": str(workbook_path),
        "artifact_hash": artifact_hash,
        "scope": scope,
        "extractor_version": EXTRACTOR_VERSION,
        "schema_version": SCHEMA_VERSION,
        "extracted_at": extracted_at.isoformat(),
        "counts": {
            "ledger_rows": len(ledger_df),
            "distinct_dates": int(ledger_df["as_of"].nunique()) if not ledger_df.empty else 0,
            "mint_rows": mint_rows,
            "issues": len(issues),
        },
    }

    ledger_df.to_parquet(out_dir / f"ledger_series_{scope}.parquet", index=False)
    (out_dir / f"extraction_manifest_{scope}.json").write_text(
        json.dumps(manifest, indent=2, default=str)
    )
    (out_dir / f"issues_{scope}.json").write_text(json.dumps(issues, indent=2, default=str))

    return {"ledger": ledger_df, "manifest": manifest, "issues": issues}
